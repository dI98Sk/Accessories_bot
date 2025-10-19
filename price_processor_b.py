import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from PriceProcessorXtremeCase import PriceProcessorXtremeCase


class PriceProcessor:
    def __init__(self, markup_value=50, price_col_index=4, header_row=5, output_subdir="processed_files"):
        """
        markup_value: сумма для добавления к цене
        price_col_index: номер колонки с ценой (1-индексация)
        header_row: строка заголовка (1-индексация)
        output_subdir: подпапка для сохранения обработанных файлов
        """
        self.markup_value = markup_value
        self.price_col_index = price_col_index
        self.header_row = header_row
        self.output_subdir = output_subdir

    def process_file(self, input_file: str, output_file: str | None = None):
        """Обрабатывает один Excel-файл, перезаписывая цены."""
        input_dir = os.path.dirname(input_file)
        output_dir = os.path.join(input_dir, self.output_subdir)
        os.makedirs(output_dir, exist_ok=True)

        if output_file is None:
            base_name = os.path.basename(input_file)
            name, ext = os.path.splitext(base_name)
            output_file = os.path.join(output_dir, f"{name}_Update{ext}")

        wb = load_workbook(input_file)
        ws = wb.active

        price_col_letter = get_column_letter(self.price_col_index)

        for row in range(self.header_row + 1, ws.max_row + 1):
            cell = ws[f"{price_col_letter}{row}"]
            try:
                price = float(cell.value)
            except (TypeError, ValueError):
                price = None

            if price is not None:
                ws[f"{price_col_letter}{row}"] = price + self.markup_value

        wb.save(output_file)
        print(f"✔ Файл сохранён: {output_file}")

    def process_directory(self, folder_path: str):
        """
        Обрабатывает все .xlsx файлы в папке, используя process_file.
        Возвращает список обработанных файлов.
        """
        processed_files = []

        for file_name in os.listdir(folder_path):
            if file_name.lower().endswith(".xlsx"):
                input_path = os.path.join(folder_path, file_name)
                try:
                    self.process_file(input_path)
                    input_dir = os.path.dirname(input_path)
                    output_file = os.path.join(input_dir, self.output_subdir, f"{os.path.splitext(file_name)[0]}_Update.xlsx")
                    processed_files.append(output_file)
                except Exception as e:
                    print(f"⚠ Не удалось обработать {file_name}: {e}")

        return processed_files


# === Пример использования ===
if __name__ == "__main__":
    processor = PriceProcessorXtremeCase(markup_value=50, price_col_index=4, header_row=5)

    # Обработка одного файла
    # processor.process_file("dataXtremeCase/Benks OPT от 14.08.2025.xlsx")

    # Обработка директории
    processor.process_directory("./dataXtremeCase")