"""
Конкретные реализации процессоров прайс-листов
"""
import os
import logging
import shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from base_processor import BasePriceProcessor
from config import ProcessorConfig
from xlsx_image_preserving_processor import XlsxImagePreservingProcessor

logger = logging.getLogger(__name__)


class StandardPriceProcessor(BasePriceProcessor):
    """
    Стандартный процессор для обработки прайс-листов.
    Обрабатывает файл целиком, добавляя наценку к ценам.
    """
    
    def _process_workbook(self, input_file: str, output_file: str) -> None:
        """
        Обрабатывает workbook стандартным способом.
        Использует специальный процессор для сохранения изображений.
        """
        # Используем процессор который работает с xlsx напрямую и сохраняет изображения
        updated_count = XlsxImagePreservingProcessor.update_prices_in_xlsx(
            input_file=input_file,
            output_file=output_file,
            markup=self.markup_value,
            price_column=self.price_col_index,
            header_row=self.header_row
        )
        
        logger.info(f"Обновлено цен: {updated_count}")
        logger.debug("Изображения и форматирование сохранены")


class MultiSheetPriceProcessor(BasePriceProcessor):
    """
    Процессор для файлов с несколькими листами.
    Разделяет каждый лист в отдельный файл.
    Используется для CifrovoyRay прайсов.
    """
    
    def _process_workbook(self, input_file: str, output_file: str) -> None:
        """
        Обрабатывает workbook с несколькими листами.
        Каждый лист сохраняется в отдельный файл.
        Сохраняет все изображения из исходного файла.
        """
        wb = load_workbook(input_file)
        output_dir = os.path.dirname(output_file)
        
        for sheet_name in wb.sheetnames:
            logger.info(f"Обработка листа: {sheet_name}")
            
            # Путь для выходного файла этого листа
            sheet_output_file = os.path.join(output_dir, f"{sheet_name}.xlsx")
            
            # Копируем исходный файл
            shutil.copy2(input_file, sheet_output_file)
            
            # Открываем для удаления лишних листов
            new_wb = load_workbook(sheet_output_file)
            
            # Удаляем все листы кроме текущего
            for other_sheet in list(new_wb.sheetnames):
                if other_sheet != sheet_name:
                    new_wb.remove(new_wb[other_sheet])
            
            # Гарантируем, что лист видим
            new_wb[sheet_name].sheet_state = "visible"
            
            # Сохраняем файл с одним листом
            new_wb.save(sheet_output_file)
            new_wb.close()
            
            # Теперь обновляем цены используя специальный процессор
            temp_output = sheet_output_file + ".tmp"
            updated_count = XlsxImagePreservingProcessor.update_prices_in_xlsx(
                input_file=sheet_output_file,
                output_file=temp_output,
                markup=self.markup_value,
                price_column=self.price_col_index,
                header_row=self.header_row
            )
            
            # Заменяем оригинал обработанным файлом
            shutil.move(temp_output, sheet_output_file)
            
            logger.info(f"✔ Лист '{sheet_name}' сохранен: {sheet_output_file} (обновлено {updated_count} цен)")
            logger.debug(f"Изображения сохранены для листа '{sheet_name}'")
        
        wb.close()
        logger.info(f"Все листы обработаны из файла: {input_file}")


class XtremeCaseProcessor(StandardPriceProcessor):
    """
    Процессор специально для XtremeCase прайсов.
    Наследуется от StandardPriceProcessor, добавляет специфическую логику при необходимости.
    """
    
    def __init__(self, config: ProcessorConfig = None):
        if config is None:
            config = ProcessorConfig.xtreme_case()
        super().__init__(config)


class CifrovoyRayProcessor(MultiSheetPriceProcessor):
    """
    Процессор специально для CifrovoyRay прайсов.
    Наследуется от MultiSheetPriceProcessor для разделения листов.
    """
    
    def __init__(self, config: ProcessorConfig = None):
        if config is None:
            config = ProcessorConfig.cifrovoy_ray()
        super().__init__(config)


class ProcessorFactory:
    """
    Фабрика для создания процессоров на основе типа файла.
    """
    
    @staticmethod
    def create_processor(file_type: str, config: ProcessorConfig = None) -> BasePriceProcessor:
        """
        Создает процессор на основе типа файла.
        
        Args:
            file_type: Тип файла ('xtreme_case', 'cifrovoy_ray', 'default')
            config: Конфигурация процессора (если None, используется стандартная для типа)
        
        Returns:
            Экземпляр процессора
        
        Raises:
            ValueError: Если тип файла неизвестен
        """
        if file_type == "xtreme_case":
            return XtremeCaseProcessor(config)
        elif file_type == "cifrovoy_ray":
            return CifrovoyRayProcessor(config)
        elif file_type == "default":
            if config is None:
                config = ProcessorConfig.default()
            return StandardPriceProcessor(config)
        else:
            raise ValueError(f"Неизвестный тип файла: {file_type}")
    
    @staticmethod
    def auto_create_processor(file_path: str) -> BasePriceProcessor:
        """
        Автоматически определяет тип файла и создает соответствующий процессор.
        
        Args:
            file_path: Путь к файлу
        
        Returns:
            Экземпляр процессора
        """
        # Создаем временный процессор для определения типа
        temp_processor = StandardPriceProcessor(ProcessorConfig.default())
        file_type = temp_processor.detect_file_type(file_path)
        
        # Создаем правильный процессор
        return ProcessorFactory.create_processor(file_type)

