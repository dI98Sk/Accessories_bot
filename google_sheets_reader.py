"""
Google Sheets Reader для чтения прайс-листов из Google Таблиц
"""
import os
import logging
import asyncio
from pathlib import Path
from typing import Optional, Callable, Awaitable, List, Dict
from datetime import datetime
import io

import gspread
from google.oauth2.service_account import Credentials
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from config import GoogleSheetsConfig, AppConfig
from logger import get_logger

logger = get_logger(__name__)


class GoogleSheetsReader:
    """
    Клиент для чтения Google Sheets и конвертации в Excel файлы.
    """
    
    # Области доступа для Google Sheets API
    SCOPES = [
        'https://www.googleapis.com/auth/spreadsheets.readonly',
        'https://www.googleapis.com/auth/drive.readonly'
    ]
    
    def __init__(self, sheets_config: GoogleSheetsConfig, app_config: AppConfig):
        """
        Args:
            sheets_config: Конфигурация Google Sheets
            app_config: Конфигурация приложения
        """
        self.sheets_config = sheets_config
        self.app_config = app_config
        
        # Клиент gspread
        self.client: Optional[gspread.Client] = None
        self.spreadsheet: Optional[gspread.Spreadsheet] = None
        
        # Обработчик новых файлов
        self.file_handler: Optional[Callable[[str, str], Awaitable[None]]] = None
        
        # Последнее обновление таблицы
        self.last_update_time: Optional[datetime] = None
        
        logger.info("GoogleSheetsReader инициализирован")
    
    def connect(self):
        """Подключается к Google Sheets API"""
        try:
            logger.info("Подключение к Google Sheets API...")
            
            # Загружаем credentials
            credentials = Credentials.from_service_account_file(
                self.sheets_config.credentials_file,
                scopes=self.SCOPES
            )
            
            # Создаем клиент
            self.client = gspread.authorize(credentials)
            
            # Открываем таблицу
            self.spreadsheet = self.client.open_by_key(self.sheets_config.spreadsheet_id)
            
            logger.info(f"✓ Подключено к таблице: {self.spreadsheet.title}")
            logger.info(f"  Количество листов: {len(self.spreadsheet.worksheets())}")
            
        except FileNotFoundError:
            logger.error(f"Файл credentials не найден: {self.sheets_config.credentials_file}")
            raise
        except Exception as e:
            logger.error(f"Ошибка при подключении к Google Sheets: {e}", exc_info=True)
            raise
    
    def set_file_handler(self, handler: Callable[[str, str], Awaitable[None]]):
        """
        Устанавливает обработчик для новых файлов.
        
        Args:
            handler: Async функция, принимающая (file_path, file_name)
        """
        self.file_handler = handler
        logger.info("Обработчик файлов установлен")
    
    async def check_for_updates(self):
        """Проверяет обновления в таблице"""
        try:
            if not self.spreadsheet:
                logger.warning("Таблица не подключена")
                return
            
            # Получаем информацию о последнем обновлении
            # Note: gspread не предоставляет прямой способ получить дату обновления
            # Будем проверять на каждом интервале
            
            logger.info("Проверка обновлений в Google Sheets...")
            
            # Экспортируем таблицу в Excel
            file_path = await self._export_to_excel()
            
            if file_path and self.file_handler:
                file_name = os.path.basename(file_path)
                try:
                    await self.file_handler(file_path, file_name)
                    logger.info(f"✓ Таблица обработана: {file_name}")
                except Exception as e:
                    logger.error(f"Ошибка при обработке таблицы: {e}")
            
            self.last_update_time = datetime.now()
        
        except Exception as e:
            logger.error(f"Ошибка при проверке обновлений: {e}", exc_info=True)
    
    async def _export_to_excel(self) -> Optional[str]:
        """
        Экспортирует Google Sheets в Excel файл.
        
        Returns:
            Путь к созданному Excel файлу или None при ошибке
        """
        try:
            # Создаем новый Excel workbook
            wb = Workbook()
            
            # Удаляем стандартный лист
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Получаем все листы из Google Sheets
            worksheets = self.spreadsheet.worksheets()
            logger.info(f"Экспорт {len(worksheets)} листов...")
            
            for gs_worksheet in worksheets:
                logger.info(f"  Экспорт листа: {gs_worksheet.title}")
                
                # Получаем все значения
                values = gs_worksheet.get_all_values()
                
                if not values:
                    logger.warning(f"  Лист {gs_worksheet.title} пустой, пропускаем")
                    continue
                
                # Создаем новый лист в Excel
                ws = wb.create_sheet(title=gs_worksheet.title)
                
                # Заполняем данными
                for row_idx, row_data in enumerate(values, start=1):
                    for col_idx, cell_value in enumerate(row_data, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=cell_value)
                
                logger.info(f"  ✓ Лист {gs_worksheet.title}: {len(values)} строк")
            
            # Сохраняем файл
            temp_dir = Path(self.app_config.temp_dir)
            temp_dir.mkdir(parents=True, exist_ok=True)
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            file_name = f"CifrovoyRay_{timestamp}.xlsx"
            file_path = temp_dir / file_name
            
            wb.save(str(file_path))
            logger.info(f"✓ Excel файл создан: {file_path}")
            
            return str(file_path)
        
        except Exception as e:
            logger.error(f"Ошибка при экспорте в Excel: {e}", exc_info=True)
            return None
    
    def get_worksheet_data(self, sheet_name: str) -> List[List[str]]:
        """
        Получает данные конкретного листа.
        
        Args:
            sheet_name: Название листа
        
        Returns:
            Список строк (каждая строка - список значений)
        """
        try:
            worksheet = self.spreadsheet.worksheet(sheet_name)
            return worksheet.get_all_values()
        except Exception as e:
            logger.error(f"Ошибка при получении данных листа {sheet_name}: {e}")
            return []
    
    async def start_monitoring(self):
        """
        Запускает мониторинг Google Sheets в бесконечном цикле.
        Проверяет обновления с заданным интервалом.
        """
        logger.info("Запуск мониторинга Google Sheets...")
        logger.info(f"Интервал проверки: {self.app_config.sheets_check_interval} секунд")
        logger.info(f"Таблица: {self.spreadsheet.title if self.spreadsheet else 'не подключена'}")
        
        while True:
            try:
                await self.check_for_updates()
                await asyncio.sleep(self.app_config.sheets_check_interval)
            except KeyboardInterrupt:
                logger.info("Мониторинг остановлен пользователем")
                break
            except Exception as e:
                logger.error(f"Ошибка в цикле мониторинга: {e}", exc_info=True)
                await asyncio.sleep(30)  # Ждем перед повтором


def test_connection():
    """Тестирует подключение к Google Sheets"""
    from config import get_google_sheets_config, get_app_config
    from logger import setup_logging
    
    app_config = get_app_config()
    setup_logging(app_config)
    
    sheets_config = get_google_sheets_config()
    
    reader = GoogleSheetsReader(sheets_config, app_config)
    reader.connect()
    
    # Выводим информацию о таблице
    logger.info("=" * 60)
    logger.info(f"Название таблицы: {reader.spreadsheet.title}")
    logger.info(f"URL: {reader.spreadsheet.url}")
    logger.info(f"Листы:")
    
    for worksheet in reader.spreadsheet.worksheets():
        row_count = worksheet.row_count
        col_count = worksheet.col_count
        logger.info(f"  - {worksheet.title}: {row_count} строк, {col_count} колонок")
    
    logger.info("=" * 60)
    
    # Тест экспорта
    async def test_export():
        file_path = await reader._export_to_excel()
        if file_path:
            logger.info(f"✓ Тестовый экспорт успешен: {file_path}")
        else:
            logger.error("✗ Ошибка при тестовом экспорте")
    
    asyncio.run(test_export())


if __name__ == "__main__":
    test_connection()

