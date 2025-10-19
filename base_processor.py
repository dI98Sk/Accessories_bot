"""
Базовый класс для обработки прайс-листов
"""
import os
import logging
from abc import ABC, abstractmethod
from typing import List, Optional
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from config import ProcessorConfig

logger = logging.getLogger(__name__)


class BasePriceProcessor(ABC):
    """
    Базовый класс для обработки прайс-листов.
    Все конкретные процессоры должны наследоваться от этого класса.
    """
    
    def __init__(self, config: ProcessorConfig):
        """
        Args:
            config: Конфигурация процессора
        """
        self.config = config
        self.markup_value = config.markup_value
        self.price_col_index = config.price_col_index
        self.header_row = config.header_row
        self.output_subdir = config.output_subdir
        logger.info(f"Инициализирован {self.__class__.__name__} с наценкой {self.markup_value}")
    
    def process_file(self, input_file: str, output_file: Optional[str] = None) -> str:
        """
        Обрабатывает один Excel-файл.
        
        Args:
            input_file: Путь к входному файлу
            output_file: Путь к выходному файлу (если None, генерируется автоматически)
        
        Returns:
            Путь к обработанному файлу
        
        Raises:
            FileNotFoundError: Если входной файл не найден
            Exception: При ошибках обработки
        """
        if not os.path.exists(input_file):
            logger.error(f"Файл не найден: {input_file}")
            raise FileNotFoundError(f"Файл не найден: {input_file}")
        
        try:
            logger.info(f"Начало обработки файла: {input_file}")
            
            # Подготовка выходной директории
            input_dir = os.path.dirname(input_file)
            output_dir = os.path.join(input_dir, self.output_subdir)
            os.makedirs(output_dir, exist_ok=True)
            
            # Генерация имени выходного файла
            if output_file is None:
                base_name = os.path.basename(input_file)
                name, ext = os.path.splitext(base_name)
                output_file = os.path.join(output_dir, f"{name}_Update{ext}")
            
            # Обработка файла (делегируем конкретной реализации)
            self._process_workbook(input_file, output_file)
            
            logger.info(f"✔ Файл успешно обработан: {output_file}")
            return output_file
            
        except Exception as e:
            logger.error(f"Ошибка при обработке {input_file}: {e}", exc_info=True)
            raise
    
    @abstractmethod
    def _process_workbook(self, input_file: str, output_file: str) -> None:
        """
        Абстрактный метод для обработки workbook.
        Должен быть реализован в дочерних классах.
        
        Args:
            input_file: Путь к входному файлу
            output_file: Путь к выходному файлу
        """
        pass
    
    def _update_prices(self, ws, start_row: Optional[int] = None) -> int:
        """
        Обновляет цены в листе Excel.
        
        Args:
            ws: Worksheet объект
            start_row: Начальная строка (если None, используется header_row + 1)
        
        Returns:
            Количество обновленных ячеек
        """
        if start_row is None:
            start_row = self.header_row + 1
        
        price_col_letter = get_column_letter(self.price_col_index)
        updated_count = 0
        
        for row in range(start_row, ws.max_row + 1):
            cell = ws[f"{price_col_letter}{row}"]
            try:
                price = float(cell.value)
            except (TypeError, ValueError):
                price = None
            
            if price is not None:
                ws[f"{price_col_letter}{row}"] = price + self.markup_value
                updated_count += 1
        
        logger.debug(f"Обновлено {updated_count} цен в листе '{ws.title}'")
        return updated_count
    
    def process_directory(self, folder_path: str) -> List[str]:
        """
        Обрабатывает все .xlsx файлы в папке.
        
        Args:
            folder_path: Путь к папке с файлами
        
        Returns:
            Список путей к обработанным файлам
        """
        if not os.path.exists(folder_path):
            logger.error(f"Директория не найдена: {folder_path}")
            return []
        
        processed_files = []
        logger.info(f"Начало обработки директории: {folder_path}")
        
        for file_name in os.listdir(folder_path):
            if file_name.lower().endswith(".xlsx") and not file_name.startswith("~"):
                input_path = os.path.join(folder_path, file_name)
                try:
                    output_path = self.process_file(input_path)
                    processed_files.append(output_path)
                except Exception as e:
                    logger.error(f"⚠ Не удалось обработать {file_name}: {e}")
        
        logger.info(f"Обработано файлов: {len(processed_files)}")
        return processed_files
    
    def detect_file_type(self, file_path: str) -> str:
        """
        Определяет тип прайс-листа по содержимому или имени файла.
        
        Args:
            file_path: Путь к файлу
        
        Returns:
            Тип файла (строка)
        """
        file_name = os.path.basename(file_path).lower()
        
        # Определение по имени файла
        if any(keyword in file_name for keyword in ["benks", "energea", "pitaka", "uag", "uniq", "vaja"]):
            return "xtreme_case"
        elif any(keyword in file_name for keyword in ["cifrovoy", "digital", "cr"]):
            return "cifrovoy_ray"
        
        # Если не удалось определить, возвращаем default
        logger.warning(f"Не удалось определить тип файла: {file_name}, используем default")
        return "default"

