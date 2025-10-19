import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class PriceProcessorCR:
    def __init__(self, markup_value=50, price_col_index=4, header_row=5, output_subdir="processed_files"):
        """
        markup_value: сумма для добавления к цене
        price_col_index: номер колонки с ценой (1-индексация, A=1, B=2, ...)
        header_row: строка заголовка (1-индексация)
        output_subdir: подпапка для сохранения обработанных файлов
        """
        self.markup_value = markup_value
        self.price_col_index = price_col_index
        self.header_row = header_row
        self.output_subdir = output_subdir

    def process_file(self, input_file: str):
        """
        Обрабатывает Excel-файл с несколькими листами.
        Каждый лист сохраняется в отдельный .xlsx файл с сохранением картинок и форматирования.
        """
        input_dir = os.path.dirname(input_file)
        output_dir = os.path.join(input_dir, self.output_subdir)
        os.makedirs(output_dir, exist_ok=True)

        wb = load_workbook(input_file)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # обработка цен
            price_col_letter = get_column_letter(self.price_col_index)
            for row in range(self.header_row + 1, ws.max_row + 1):
                cell = ws[f"{price_col_letter}{row}"]
                try:
                    price = float(cell.value)
                except (TypeError, ValueError):
                    price = None
                if price is not None:
                    ws[f"{price_col_letter}{row}"] = price + self.markup_value

            # создаём копию файла и оставляем только нужный лист
            output_file = os.path.join(output_dir, f"{sheet_name}.xlsx")
            new_wb = load_workbook(input_file)  # читаем исходный файл заново (с картинками)

            # удаляем все листы кроме текущего
            for other in list(new_wb.sheetnames):
                if other != sheet_name:
                    new_wb.remove(new_wb[other])

            # гарантируем, что лист не скрыт
            new_wb[sheet_name].sheet_state = "visible"

            # обновляем цены в копии
            new_ws = new_wb[sheet_name]
            price_col_letter = get_column_letter(self.price_col_index)
            for row in range(self.header_row + 1, new_ws.max_row + 1):
                cell = new_ws[f"{price_col_letter}{row}"]
                try:
                    price = float(cell.value)
                except (TypeError, ValueError):
                    price = None
                if price is not None:
                    new_ws[f"{price_col_letter}{row}"] = price + self.markup_value

            # сохраняем
            new_wb.save(output_file)
            print(f"✔ Лист '{sheet_name}' сохранён: {output_file}")

        print(f"=== Все листы обработаны и сохранены в папку: {output_dir} ===")

    def process_directory(self, folder_path: str):
        """
        Обрабатывает все .xlsx файлы в папке.
        """
        processed_files = []
        for file_name in os.listdir(folder_path):
            if file_name.lower().endswith(".xlsx"):
                input_path = os.path.join(folder_path, file_name)
                try:
                    self.process_file(input_path)
                    processed_files.append(input_path)
                except Exception as e:
                    print(f"⚠ Не удалось обработать {file_name}: {e}")
        return processed_files


# === Пример использования ===
if __name__ == "__main__":
    processor = PriceProcessorCR(markup_value=50, price_col_index=4, header_row=5)

    # Обработка одного файла
    processor.process_file("./dataCifrovoyRay/12.09 Прайс Лист CR.xlsx")

    # Обработка директории
    # processor.process_directory("./dataCifrovoyRay")